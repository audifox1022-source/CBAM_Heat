import streamlit as st
import pandas as pd
import re
import io
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

# -----------------------------------------------------------
# Streamlit 페이지 설정
# -----------------------------------------------------------
st.set_page_config(page_title="CBAM 데이터 통합기", page_icon="🏭", layout="wide")

st.title("🏭 열처리 작업지시서 통합 도구 (Pro)")
st.markdown("""
**CSV 및 Excel 파일**을 업로드하면 깔끔한 보고서 형태로 합쳐줍니다.
* **(숫자)** 호기로 표시된 파일만 통합합니다.
* **(단조)** 파일은 자동으로 제외합니다.
* **컬럼명을 자동으로 통일**하여 데이터 누락을 방지합니다.
""")

def read_csv_with_encoding(file_obj, **kwargs):
    """CSV 파일을 읽을 때 인코딩(utf-8, cp949 등) 자동 감지"""
    encodings = ['utf-8', 'cp949', 'euc-kr']
    for enc in encodings:
        try:
            file_obj.seek(0)
            return pd.read_csv(file_obj, encoding=enc, **kwargs)
        except:
            continue
    file_obj.seek(0)
    return pd.read_csv(file_obj, encoding='utf-8', **kwargs)

def find_header_row(file_obj, file_ext):
    """
    실제 데이터 헤더가 있는 행 번호를 찾습니다.
    (수주NO, 품명, 수량 등의 키워드가 많이 포함된 행을 헤더로 판단)
    """
    try:
        file_obj.seek(0)
        if file_ext == '.csv':
            df_temp = read_csv_with_encoding(file_obj, header=None, nrows=20)
        else:
            df_temp = pd.read_excel(file_obj, header=None, nrows=20)

        # 헤더로 의심되는 키워드 목록
        keywords = ['수주', 'NO', '품명', '품 명', '규격', '재질', '중량']
        
        max_score = 0
        best_row = 0

        for i, row in df_temp.iterrows():
            row_str = row.astype(str).values
            # 해당 행에 키워드가 몇 개나 포함되어 있는지 점수 매기기
            score = sum(1 for keyword in keywords if any(keyword in str(cell) for cell in row_str))
            
            if score > max_score:
                max_score = score
                best_row = i
        
        # 키워드가 2개 이상 발견된 행을 헤더로 인정, 아니면 0번째 줄
        return best_row if max_score >= 2 else 0
        
    except Exception as e:
        return 0

def clean_column_names(df):
    """
    컬럼명을 표준화하여 데이터가 흩어지는 것을 방지합니다.
    예: '품 명' -> '품명', '수주NO.' -> '수주NO'
    """
    # 1. 공백 제거 및 특수문자 정리
    df.columns = df.columns.astype(str).str.replace(' ', '').str.replace('.', '').str.replace('\n', '')
    
    # 2. 유사한 컬럼명 통일 (매핑 테이블)
    rename_map = {
        '수주번호': '수주NO',
        '지시서번호': '지시서NO',
        '지시번호': '지시서NO',
        '품목': '품명',
        '재질': '재질',
        '원소재': '재질'
    }
    
    # 컬럼명 변경 적용
    new_columns = {}
    for col in df.columns:
        for key, value in rename_map.items():
            if key in col:
                new_columns[col] = value
                break
    
    if new_columns:
        df = df.rename(columns=new_columns)
        
    return df

def style_excel(writer, df):
    """엑셀 파일에 테두리, 배경색, 열 너비 자동 맞춤 적용"""
    workbook = writer.book
    worksheet = writer.sheets['Sheet1']
    
    # 스타일 정의
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    border_style = Border(left=Side(style='thin'), right=Side(style='thin'), 
                          top=Side(style='thin'), bottom=Side(style='thin'))
    center_align = Alignment(horizontal='center', vertical='center')

    # 1. 헤더 스타일 적용
    for col_num, value in enumerate(df.columns.values):
        cell = worksheet.cell(row=1, column=col_num + 1)
        cell.value = value
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = border_style

    # 2. 데이터 스타일 적용 및 열 너비 자동 조정
    for i, col in enumerate(df.columns):
        max_length = 0
        column = col
        
        # 헤더 길이 측정
        try:
            if len(str(column)) > max_length:
                max_length = len(str(column))
        except:
            pass

        # 데이터 길이 측정 (상위 100개만 샘플링하여 속도 향상)
        for cell in worksheet[get_column_letter(i+1)][1:101]: 
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
                cell.border = border_style # 테두리 적용
            except:
                pass
        
        # 열 너비 설정 (최대 50으로 제한)
        adjusted_width = (max_length + 2)
        if adjusted_width > 50:
            adjusted_width = 50
        worksheet.column_dimensions[get_column_letter(i+1)].width = adjusted_width

# -----------------------------------------------------------
# 메인 로직
# -----------------------------------------------------------
uploaded_files = st.file_uploader(
    "여기에 파일을 드래그하세요 (CSV, XLSX, XLS)", 
    accept_multiple_files=True, 
    type=['csv', 'xlsx', 'xls']
)

if uploaded_files:
    if st.button("데이터 통합 시작"):
        master_df = pd.DataFrame()
        success_count = 0
        skip_count = 0
        error_log = []
        
        progress_bar = st.progress(0)
        status_text = st.empty()

        for idx, uploaded_file in enumerate(uploaded_files):
            try:
                filename = uploaded_file.name
                file_ext = os.path.splitext(filename)[1].lower()
                
                # (1) 파일명 분석
                date_match = re.search(r"(\d{1,2}-\d{1,2})", filename)
                furnace_match = re.search(r"\((.+?)\)", filename)

                work_date = date_match.group(1) if date_match else "날짜미상"
                furnace_no = furnace_match.group(1) if furnace_match else "호기미상"

                # 단조 파일 필터링
                if "단조" in furnace_no:
                    status_text.text(f"⛔ 제외됨 (단조): {filename}")
                    skip_count += 1
                    progress_bar.progress((idx + 1) / len(uploaded_files))
                    continue

                status_text.text(f"🔄 처리 중: {filename}")

                # (2) 헤더 찾기 및 읽기
                header_idx = find_header_row(uploaded_file, file_ext)
                uploaded_file.seek(0)

                if file_ext == '.csv':
                    df = read_csv_with_encoding(uploaded_file, header=header_idx)
                else:
                    df = pd.read_excel(uploaded_file, header=header_idx)

                # (3) 컬럼명 표준화 (데이터 누락 방지 핵심!)
                df = clean_column_names(df)

                # (4) 유효 데이터 필터링 ('수주NO' 컬럼이 있는 경우만)
                target_cols = [c for c in df.columns if "수주" in str(c) or "NO" in str(c)]
                
                if target_cols:
                    # '수주NO' 또는 'NO' 컬럼이 비어있지 않은 행만 선택
                    valid_rows = df[df[target_cols[0]].notna()].copy()
                    
                    if not valid_rows.empty:
                        # 메타데이터 추가
                        valid_rows.insert(0, '지시서번호(호기)', furnace_no)
                        valid_rows.insert(0, '작업지시일', work_date)
                        
                        master_df = pd.concat([master_df, valid_rows], ignore_index=True)
                        success_count += 1
                    else:
                        # 데이터는 없지만 파일은 정상인 경우 (빈 양식 등)
                        pass
                else:
                    error_log.append(f"⚠️ {filename}: 유효한 데이터 헤더를 찾을 수 없음")
                
            except Exception as e:
                error_log.append(f"❌ {filename}: {str(e)}")
            
            progress_bar.progress((idx + 1) / len(uploaded_files))

        status_text.text("모든 작업 완료!")

        # -----------------------------------------------------------
        # 결과 출력 및 다운로드
        # -----------------------------------------------------------
        if not master_df.empty:
            st.success(f"✅ 통합 완료! (총 {success_count}개 파일 합침, {skip_count}개 단조 파일 제외)")
            
            if error_log:
                st.warning(f"⚠️ {len(error_log)}개 파일 처리 중 문제 발생 (나머지는 정상 통합됨)")
                with st.expander("문제 발생 로그 보기"):
                    for err in error_log:
                        st.write(err)

            # 데이터 미리보기
            st.subheader("📊 통합 데이터 미리보기")
            st.dataframe(master_df.head())

            # 엑셀 다운로드 (서식 적용)
            buffer = io.BytesIO()
            with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
                master_df.to_excel(writer, index=False, sheet_name='Sheet1')
                style_excel(writer, master_df) # 서식 적용 함수 호출
            
            st.download_button(
                label="📥 깔끔한 엑셀 파일 다운로드 (Click)",
                data=buffer,
                file_name="통합_RAW_DATA_결과.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        else:
            if skip_count > 0:
                st.warning("단조 파일을 제외하니 통합할 유효 데이터가 없습니다.")
            else:
                st.error("통합할 데이터가 없습니다. 파일을 확인해주세요.")
